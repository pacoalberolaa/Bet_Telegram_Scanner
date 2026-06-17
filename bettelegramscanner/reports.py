"""Exportación a Excel: resumen, picks verificados y picks no_verificables."""
from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .analytics import TipsterReport
from .models import BasketballLeg, DartsLeg, FootballLeg, PickDocument, TennisLeg

log = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")


def _write_header(ws, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _format_linea(linea: float | None) -> str:
    if linea is None:
        return ""
    if linea == int(linea):
        return f"{linea:+g}"
    return f"{linea:+g}"


def _leg_detail(leg) -> str:
    """Texto humano de una pierna: 'EquipoA vs EquipoB | Competicion | mercado: seleccion linea'."""
    if isinstance(leg, (FootballLeg, BasketballLeg)):
        partido = f"{leg.equipo_local} vs {leg.equipo_visitante}"
        comp = leg.competicion or ""
    elif isinstance(leg, (TennisLeg, DartsLeg)):
        partido = f"{leg.jugador_1} vs {leg.jugador_2}"
        comp = getattr(leg, "torneo", None) or getattr(leg, "competicion", None) or ""
    else:
        partido = ""
        comp = ""
    mercado = f"{leg.mercado}: {leg.seleccion}"
    if isinstance(leg, BasketballLeg) and leg.over_under is not None:
        mercado += f" [{leg.over_under}]"
    if leg.linea is not None:
        mercado += f" {_format_linea(leg.linea)}"
    parts = [partido]
    if comp:
        parts.append(comp)
    parts.append(mercado)
    return " | ".join(p for p in parts if p)


def _pick_detail(pick: PickDocument) -> str:
    """Resumen humano de todas las piernas, separadas por ' || '."""
    if not pick.payload.legs:
        return "(boleto sin piernas)"
    return " || ".join(_leg_detail(l) for l in pick.payload.legs)


def _legs_summary(pick: PickDocument) -> str:
    parts = []
    for leg in pick.payload.legs:
        if isinstance(leg, TennisLeg):
            line = f"{leg.jugador_1} vs {leg.jugador_2} | {leg.mercado}:{leg.seleccion}"
        elif isinstance(leg, FootballLeg):
            line = f"{leg.equipo_local} vs {leg.equipo_visitante} | {leg.mercado}:{leg.seleccion}"
        elif isinstance(leg, DartsLeg):
            line = f"{leg.jugador_1} vs {leg.jugador_2} | {leg.mercado}:{leg.seleccion}"
        elif isinstance(leg, BasketballLeg):
            line = f"{leg.equipo_local} vs {leg.equipo_visitante} | {leg.mercado}:{leg.seleccion}"
            if leg.over_under is not None:
                line += f" [{leg.over_under}]"
        else:
            line = str(leg)
        if leg.linea is not None:
            line += f" ({leg.linea:+g})"
        parts.append(line)
    return " || ".join(parts)


def _real_score(pick: PickDocument) -> str:
    if pick.resolution is None:
        return ""
    return " || ".join(r.marcador or "" for r in pick.resolution.legs)


def _unverif_reason(pick: PickDocument) -> str:
    if pick.resolution is None:
        return "sin resolución"
    parts = [pick.resolution.motivo or ""]
    for i, r in enumerate(pick.resolution.legs):
        if r.status == "no_verificable":
            parts.append(f"leg{i+1}: {r.motivo or '?'}")
    return " | ".join(p for p in parts if p)


def export_xlsx(out_path: Path, report: TipsterReport, picks: list[PickDocument]) -> Path:
    wb = Workbook()

    # Resumen
    ws = wb.active
    ws.title = "Resumen"
    _write_header(ws, ["Métrica", "Valor"])
    for i, (k, v) in enumerate([
        ("tipster", report.tipster),
        ("total_picks", report.total_picks),
        ("verificados", report.verificados),
        ("no_verificables", report.no_verificables),
        ("ganados", report.ganados),
        ("perdidos", report.perdidos),
        ("voids", report.voids),
        ("stake_total (u)", report.stake_total),
        ("profit_total (u)", report.profit_total),
        ("yield (%)", report.yield_pct),
    ], start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    # Picks verificados (solo los reales)
    ws_v = wb.create_sheet("Picks_Verificados")
    _write_header(ws_v, [
        "fecha_utc", "message_id", "sport", "casa", "legs", "cuota_total",
        "stake_u", "resultado_real", "marcador_real", "profit_u",
    ])
    row = 2
    for p in picks:
        if p.resolution is None or p.resolution.status == "no_verificable":
            continue
        ws_v.cell(row=row, column=1, value=p.date_utc.isoformat(sep=" ", timespec="minutes"))
        ws_v.cell(row=row, column=2, value=p.message_id)
        ws_v.cell(row=row, column=3, value=p.payload.sport)
        ws_v.cell(row=row, column=4, value=p.payload.casa_apuestas)
        ws_v.cell(row=row, column=5, value=_legs_summary(p))
        ws_v.cell(row=row, column=6, value=p.payload.cuota_total)
        ws_v.cell(row=row, column=7, value=p.payload.stake_indicado)
        ws_v.cell(row=row, column=8, value=p.resolution.status)
        ws_v.cell(row=row, column=9, value=_real_score(p))
        ws_v.cell(row=row, column=10, value=p.profit_units)
        row += 1
    for col, width in zip("ABCDEFGHIJ", (18, 12, 12, 14, 70, 12, 10, 14, 22, 12)):
        ws_v.column_dimensions[col].width = width

    # Picks no verificables (los que necesitan revisión manual)
    ws_n = wb.create_sheet("Picks_No_Verificables")
    _write_header(ws_n, [
        "fecha_utc", "message_id", "sport", "casa", "legs", "cuota_total",
        "stake_u", "motivo",
    ])
    row = 2
    for p in picks:
        if p.resolution is not None and p.resolution.status != "no_verificable":
            continue
        ws_n.cell(row=row, column=1, value=p.date_utc.isoformat(sep=" ", timespec="minutes"))
        ws_n.cell(row=row, column=2, value=p.message_id)
        ws_n.cell(row=row, column=3, value=p.payload.sport)
        ws_n.cell(row=row, column=4, value=p.payload.casa_apuestas)
        ws_n.cell(row=row, column=5, value=_legs_summary(p))
        ws_n.cell(row=row, column=6, value=p.payload.cuota_total)
        ws_n.cell(row=row, column=7, value=p.payload.stake_indicado)
        ws_n.cell(row=row, column=8, value=_unverif_reason(p))
        row += 1
    for col, width in zip("ABCDEFGH", (18, 12, 12, 14, 70, 12, 10, 60)):
        ws_n.column_dimensions[col].width = width

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Excel generado en %s", out_path)
    return out_path


def export_unverified_xlsx(out_path: Path, picks: list[PickDocument]) -> Path | None:
    """Genera un Excel SOLO con los picks no_verificables, con detalle humano.
    Devuelve la ruta o None si no había no_verificables."""
    unverified = [
        p for p in picks
        if p.resolution is None or p.resolution.status == "no_verificable"
    ]
    if not unverified:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Sin_Resultado"
    _write_header(ws, [
        "fecha", "message_id", "sport", "detalle", "cuota", "stake_u", "motivo",
    ])
    row = 2
    for p in sorted(unverified, key=lambda x: x.date_utc):
        ws.cell(row=row, column=1, value=p.date_utc.date().isoformat())
        ws.cell(row=row, column=2, value=p.message_id)
        ws.cell(row=row, column=3, value=p.payload.sport)
        ws.cell(row=row, column=4, value=_pick_detail(p))
        ws.cell(row=row, column=5, value=p.payload.cuota_total)
        ws.cell(row=row, column=6, value=p.payload.stake_indicado)
        ws.cell(row=row, column=7, value=_unverif_reason(p))
        row += 1
    for col, width in zip("ABCDEFG", (12, 12, 12, 80, 8, 10, 70)):
        ws.column_dimensions[col].width = width

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Excel de no_verificables generado en %s", out_path)
    return out_path
